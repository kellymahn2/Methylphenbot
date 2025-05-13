import logging
from telegram import Update, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ConversationHandler, ContextTypes
import openpyxl
from openpyxl import Workbook
import os
import re
from datetime import datetime

# تنظیمات لاگ
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

# مراحل گفتگو
CHOOSING, QUESTION, QUESTION1, ANSWER = range(4)

# سوالی که می‌خواهید بپرسید
SURVEY_QUESTIONS = ["موجودی داروهای زیر را به عدد قرص/کپسول وارد نمایید\n" \
                    "ریتالین۱۰\n" \
                    "روبیفن ۱۰\n" \
                    "ریتافن ۱۰\n"\
                    "متیل فنیدیت ۱۸\n"\
                    "متیل فنیدیت ۳۶\n"\
                    "متیل فنیدیت ۵۴\n",
                    "موجودی داروهای زیر را به عدد قرص/کپسول وارد نمایید\n"\
                    "ادواگرف نیم\n"\
                    "ادواگرف یک\n"\
                    "ادواگرف ۳\n"\
                    "ادواگرف ۵\n"\
                    "پروگرف نیم\n"\
                    "پروگرف ۱\n"\
                    "سل سپت \n"\
                    "مایفورتیک\n"\
                    "ایمینورال ۲۵\n"\
                    "ایمینورال ۵۰\n"\
                    "نئورال ۲۵\n"\
                    "نئورال ۵۰\n"\
                    "شربت نئورال\n"\
                    "سرتیکان\n"
                    ]

SURVEY_QUESTIONS_LINE_COUNT = [6, 14]

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """شروع گفتگو و پرسش سوال."""
    user = update.message.from_user
    logger.info("User %s started the conversation.", user.first_name)
    
    await update.message.reply_text(
        f"سلام {user.first_name}! 👋\n\n"\
        "لطفا یک دسته سوال را انتخاب کنید.\n"\
        " سوال اول: /q0\n"\
        "سوال دوم: /q1\n",
        reply_markup=ReplyKeyboardRemove(),
    )
    return QUESTION

def save_to_excel(question_index: int, user_data: dict):
    """ذخیره پاسخ‌ها در فایل اکسل"""
    filename = "survey_responses.xlsx"
    
    # اگر فایل وجود ندارد، یک فایل جدید ایجاد کن
    if not os.path.exists(filename):
        wb = Workbook() 
        wb.active.title = "Survey Responses 1"

        wb.create_sheet("Survey Responses 2", 1)

        ws = wb.worksheets[0]

        headers = ["User ID", "First Name", "Last Name", "Username"]
        headers.extend(f"Response{i + 1}" for i in range(SURVEY_QUESTIONS_LINE_COUNT[0]))
        headers.append("Timestamp")
        ws.append(headers)

        ws = wb.worksheets[1]

        headers = ["User ID", "First Name", "Last Name", "Username"]
        headers.extend(f"Response{i + 1}" for i in range(SURVEY_QUESTIONS_LINE_COUNT[1]))
        headers.append("Timestamp")
        ws.append(headers)

    else:
        wb = openpyxl.load_workbook(filename)
    
    values = [user_data['user_id'],
        user_data['first_name'],
        user_data.get('last_name', ''),
        user_data.get('username', '')]
    values.extend(user_data['response-columns'])
    values.append(user_data['timestamp'])
    
    # اضافه کردن داده جدید
    wb.worksheets[question_index].append(values)
    
    # ذخیره فایل
    wb.save(filename)


def verify_answer(question_index: int, response: str | None) -> bool:
    if(response is None):
        return False
    if(question_index not in range(2)):
        return False
    line_count = SURVEY_QUESTIONS_LINE_COUNT[question_index]
    lines = list(map(lambda x: x.strip(), response.strip().splitlines()))

    if(len(lines) != line_count):
        return False
    def is_numeric(s):
        try:
            float(s)
            return True
        except ValueError:
            return False

    return all(is_numeric(line) for line in lines)

async def receive_answer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """دریافت پاسخ کاربر و ذخیره آن"""
    user = update.message.from_user
    response = update.message.text
    question_index = context.user_data['question_index']

    if(not verify_answer(question_index, response)):
        await update.message.reply_text(
        "لطفا در هر خط موجودی همان دارو را بنویسید",
        reply_markup=ReplyKeyboardRemove()
        )
        return ANSWER

    user_data = {
        'user_id': user.id,
        'first_name': user.first_name,
        'last_name': user.last_name if user.last_name else '',
        'username': user.username if user.username else '',
        'response-columns': list(map(lambda x: x.strip(), response.strip().splitlines())),
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    save_to_excel(question_index, user_data)
    
    logger.info("Response from %s: %s", user.first_name, response)
    await update.message.reply_text(
        "ممنون از پاسخ شما! 🙏\n"
        "پاسخ شما با موفقیت ثبت شد.",
        reply_markup=ReplyKeyboardRemove()
    )
    
    return QUESTION

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """لغو گفتگو توسط کاربر"""
    user = update.message.from_user
    logger.info("User %s canceled the conversation.", user.first_name)
    await update.message.reply_text(
        'نظرسنجی لغو شد. ممنون از وقتی که گذاشتید.',
        reply_markup=ReplyKeyboardRemove()
    )
    
    return ConversationHandler.END

async def send_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ارسال فایل اکسل به ادمین"""
    user = update.message.from_user
    if user.id != 209067725:  # جایگزین کنید با آیدی عددی خودتان
        await update.message.reply_text("شما مجاز به انجام این کار نیستید.")
        return
    
    filename = "survey_responses.xlsx"
    if os.path.exists(filename):
        await update.message.reply_document(document=open(filename, 'rb'))
    else:
        await update.message.reply_text("هنوز پاسخی ثبت نشده است.")

async def question_handler(index: int, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        SURVEY_QUESTIONS[index],
        reply_markup=ReplyKeyboardRemove()
    )

    context.user_data['question_index'] = index

    return ANSWER

def main():
    """راه اندازی و اجرای بات"""
    # توکن بات خود را اینجا قرار دهید
    application = Application.builder().token("8066468395:AAHe3oAKjmD727jiIpDehc5UAtaKzj7-r98").build()  # جایگزین کنید با توکن واقعی
    
    async def question_handler0(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        return await question_handler(0, update, context)
        
    
    async def question_handler1(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        return await question_handler(1, update, context)
    
    # گفتگو برای نظرسنجی
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            QUESTION: [CommandHandler('q0', question_handler0), CommandHandler('q1', question_handler1)],
            ANSWER: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_answer)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    application.add_handler(conv_handler)
    
    # دستور برای دریافت فایل اکسل
    application.add_handler(CommandHandler('getdata', send_excel))
    
    # شروع بات
    application.run_polling()

if __name__ == '__main__':
    main()